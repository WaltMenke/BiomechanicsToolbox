import numpy as np
import pandas as pd
import os
import re
import csv
import ast
import sys
from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import tkinter as tk
from tkinter import ttk, filedialog, PhotoImage, messagebox, simpledialog
import ttkbootstrap as ttk


def list_csv_files(directory):
    csv_files = [file for file in os.listdir(directory) if file.endswith(".csv")]
    csv_files.sort(
        key=lambda x: [
            int(num) if num.isdigit() else num
            for num in re.findall(r"S(\d+)_C(\d+)", x)[0]
        ]
    )
    return csv_files


def extract_subject_number(file_name):
    return file_name.split("_")[0][1:]  # Extracts the number after 'S'


def extract_condition_number(file_name):
    return file_name.split("_")[1][1:]  # Extracts the number after 'C'


def replace_nan(value):
    if np.isnan(value):
        return "NaN"
    else:
        return value


def generate_value_rows(variable_names):
    value_rows = []
    variable_count = len(variable_names)
    for _ in range(3):
        for variable in variable_names:
            for event in events:
                value_rows.append([variable, "Value", event])
    return value_rows[: variable_count * 3]


# Function to generate row names for Index
def generate_index_rows(variable_names):
    index_rows = []
    variable_count = len(variable_names)
    for _ in range(3):
        for variable in variable_names:
            for event in events:
                index_rows.append([variable, "Index", event])
    return index_rows[: variable_count * 3]


def generate_per_loc_rows(variable_names):
    per_loc_rows = []
    variable_count = len(variable_names)
    for _ in range(3):
        for variable in variable_names:
            for event in events:
                per_loc_rows.append([variable, "Per_Loc", event])
    return per_loc_rows[: variable_count * 3]


# Subprocess inputs from main window
root = ttk.Window()
root.withdraw()
root.iconbitmap("BT_Icon.ico")
root.iconbitmap(default="BT_Icon.ico")

try:
    compile_in = sys.argv[1]
except IndexError:
    print(
        "Error: This script isn't expected to run directly, please run BiomechanicsToolbox.py instead."
    )
    messagebox.showerror(
        "Error",
        "This script isn't expected to run directly, please run BiomechanicsToolbox.py instead.",
    )
    sys.exit()
compile_out = sys.argv[2]

result = messagebox.askyesno(
    "Compile",
    "Before compiling, are you sure the input directory only contains properly named CSV Event Output files?\n\nNaming schema should follow:\n'S1_C1_Maxima.csv', 'S1_C1_Minima.csv', etc.",
    icon="question",
)
if not result:
    messagebox.showinfo("Compile Canceled", "Compile operation canceled.")
    sys.exit()

num_entries = None

maxima_stats_by_condition = {}
minima_stats_by_condition = {}

reorg_max = {}
reorg_min = {}

maxima_averages = {}
maxima_std_devs = {}
minima_averages = {}
minima_std_devs = {}

events = ["Event 1", "Event 2", "Event 3"]

headers = [
    "SUBJECT",
    "VARIABLE",
    "TYPE",
    "NUMBER",
    "Maxima Avg",
    "Maxima Stdev",
    "Minima Avg",
    "Minima Stdev",
]

pattern = r"^S\d{1,2}_C\d{1,2}_(Maxima|Minima)\.csv$"  # Ensures naming structure is like 'S1_C1_Maxima.csv'
relevant_files = list_csv_files(compile_in)
for file_name in relevant_files:
    if not re.match(pattern, file_name):
        messagebox.showerror(
            "Error",
            f"File {file_name} does not match the required naming structure.",
        )
        sys.exit()
else:
    pass

ref_file = relevant_files[0]
cond_ref = ref_file.split("_")[1][1:]

present_conditions = [cond_ref]
for file in relevant_files:
    current_cond = file.split("_")[1][1:]
    if cond_ref != current_cond and current_cond not in present_conditions:
        present_conditions.append(current_cond)

sub_count = []
for file in relevant_files:
    subject = extract_subject_number(file)
    if subject not in sub_count:
        sub_count.append(subject)
files_by_condition = {cond: [] for cond in present_conditions}


for file in relevant_files:  # Splits up files by Condition number
    condition_number = extract_condition_number(file)
    files_by_condition[condition_number].append(file)


for condition, files in files_by_condition.items():
    if num_entries is None:
        num_entries = len(files)
    else:
        if len(files) != num_entries:
            messagebox.showerror(
                "Warning",
                f"Number of entries in condition {condition} differs from others. Check the file inputs in {compile_in} and try again.",
            )
            sys.exit()
else:
    pass

for condition, files in files_by_condition.items():
    maxima_count = sum(1 for file in files if "Maxima" in file)
    minima_count = sum(1 for file in files if "Minima" in file)
    if maxima_count != minima_count:
        messagebox.showerror(
            "Warning",
            f"Number of 'Maxima' files does not match number of 'Minima' files in condition {condition}. Check the file inputs in {compile_in} and try again.",
        )
        sys.exit()
    else:
        pass

all_var_reference = []
for condition, files in files_by_condition.items():
    reference_third_line = None
    for file_name in files:
        file_path = os.path.join(compile_in, file_name)
        prefix = file_name.split(".")[0]
        with open(file_path, "r", newline="") as csv_file:
            csv_reader = csv.reader(csv_file)
            first_line_str = "".join(prefix)
            prefix_without_extension = prefix.split(".")[0]
            if not first_line_str == prefix_without_extension:
                messagebox.showerror(
                    "Warning",
                    f"File {file_name} does not match expected naming structure. Check the file inputs in {compile_in} and try again.",
                )
                sys.exit()
            next(csv_reader)
            next(csv_reader)
            third_line = next(csv_reader)
            ignore_mismatch = False
            if reference_third_line is None:
                ref_file = file_name
                reference_third_line = third_line
                all_var_reference.append(reference_third_line)
            else:
                if third_line != reference_third_line and not ignore_mismatch:
                    result = messagebox.askokcancel(
                        "Warning!",
                        f"Third line of data in file {file_name} is different relative to expected from {ref_file}. Check the file inputs in {compile_in}. It is possible the CSV files do not have identical variable counts.\n\nYou may select 'Ok' to ignore this warning or 'Cancel' to terminate the operation.",
                    )
                    if not result:
                        sys.exit()
                    else:
                        ignore_mismatch = True


lists = [ast.literal_eval(entry) for entry in reference_third_line if entry]
flat_entries = [item for sublist in lists for item in sublist]
variable_names = list(OrderedDict.fromkeys(flat_entries))

for condition, files in files_by_condition.items():
    maxima_stats = []
    minima_stats = []

    for file_name in files:
        file_path = os.path.join(compile_in, file_name)
        file_type = "Maxima" if "Maxima" in file_name else "Minima"

        with open(file_path, "r", newline="") as csv_file:
            csv_reader = csv.reader(csv_file)
            next(csv_reader)

            for row in csv_reader:
                try:
                    row = [float(value) if value != "nan" else np.nan for value in row]

                    if not any(np.isnan(row)):
                        row_avg = np.nanmean(row)
                        row_std = np.nanstd(row)
                    else:
                        row_avg = np.nan
                        row_std = np.nan

                    if file_type == "Maxima":
                        maxima_stats.append((row_avg, row_std))
                    else:
                        minima_stats.append((row_avg, row_std))
                except ValueError:  # Catches non-numeric rows
                    continue

    maxima_stats_by_condition[condition] = maxima_stats
    minima_stats_by_condition[condition] = minima_stats

workbook = Workbook()
num_events = 3
rows_per_sub = 3 * 3 * len(variable_names)
for condition, max_data in maxima_stats_by_condition.items():
    sheet = workbook.create_sheet(title="C" + str(condition))
    sheet.append(headers)
    row_idx = 2

    for data_label, data_matrix, start_col in [
        ("maxima", max_data, 5),
        ("minima", minima_stats_by_condition[condition], 7),
    ]:
        for i, row_data in enumerate(data_matrix, start=row_idx):
            for j, value in enumerate(row_data, start=start_col):
                sheet.cell(row=i, column=j).value = replace_nan(value)

    all_subject_data = []
    for subject in sub_count:
        subject_data = []
        for data_type in ["Value", "Index", "Per_Loc"]:
            for variable in variable_names:
                clean_var = variable.replace("Right_", "").replace("Left_", "")
                for event_idx in range(1, num_events + 1):
                    subject_data.append(
                        [subject, clean_var, data_type, f"Event {event_idx}"]
                    )
        all_subject_data.extend(subject_data)

    rows_to_copy = []
    for row in range(2, sheet.max_row + 1):
        row_data = [
            sheet.cell(row=row, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]
        rows_to_copy.append(row_data)

    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None

    for i in range(rows_per_sub):
        for subject_idx in range(len(sub_count)):
            old_idx = subject_idx * rows_per_sub + i
            new_row = (
                i * len(sub_count) + subject_idx + 2
            )  # +2 because row 1 is headers

            if old_idx < len(all_subject_data):
                for col_idx, value in enumerate(all_subject_data[old_idx], start=1):
                    sheet.cell(row=new_row, column=col_idx).value = value

            if old_idx < len(rows_to_copy):
                row_data = rows_to_copy[old_idx]
                for col_idx in range(
                    5, len(row_data) + 1
                ):  # Start from column 5 (data columns)
                    sheet.cell(row=new_row, column=col_idx).value = row_data[
                        col_idx - 1
                    ]


workbook.remove(workbook.active)
savepath = filedialog.asksaveasfilename(
    defaultextension=".xlsx",
    filetypes=[("Excel files", "*.xlsx")],
    initialfile="Events_Individual.xlsx",
)
if savepath:
    workbook.save(filename=savepath)
else:
    messagebox.showinfo(
        "Save Error", "Save operation canceled by user. Returning to main window."
    )
    sys.exit()
workbook.close()


sheets = pd.read_excel(savepath, sheet_name=None)
output_sheets = {}
type_order = {"Value": 0, "Index": 1, "Per_Loc": 2}
for sheet_name, df in sheets.items():
    results = []
    # Iterate over unique combinations of VARIABLE, TYPE, and NUMBER
    for variable in sorted(df["VARIABLE"].unique()):  # Sort VARIABLE alphabetically
        clean_var = variable.replace("Right_", "").replace("Left_", "")
        for event in sorted(df["NUMBER"].unique()):  # Sort events numerically
            for data_type in ["Value", "Index", "Per_Loc"]:
                filtered = df[
                    (df["VARIABLE"] == clean_var)
                    & (df["TYPE"] == data_type)
                    & (df["NUMBER"] == event)
                ]
                for i in range(0, len(filtered), len(sub_count)):
                    chunk = filtered.iloc[i : i + len(sub_count)]
                    maxima_avg = chunk["Maxima Avg"].mean(skipna=True)
                    maxima_stdev = chunk["Maxima Avg"].std(skipna=True)
                    minima_avg = chunk["Minima Avg"].mean(skipna=True)
                    minima_stdev = chunk["Minima Avg"].std(skipna=True)

                    # Replace empty or NaN values
                    result_dict = {
                        "VARIABLE": clean_var,
                        "TYPE": data_type,
                        "NUMBER": event,
                        "Maxima Avg": np.nan if pd.isna(maxima_avg) else maxima_avg,
                        "Maxima Stdev": (
                            np.nan if pd.isna(maxima_stdev) else maxima_stdev
                        ),
                        "Minima Avg": np.nan if pd.isna(minima_avg) else minima_avg,
                        "Minima Stdev": (
                            np.nan if pd.isna(minima_stdev) else minima_stdev
                        ),
                    }
                    results.append(result_dict)

    processed_df = pd.DataFrame(results)

    # Fill any remaining NaN values
    processed_df = processed_df.fillna(value=np.nan)

    # Sort by TYPE (custom order), then by VARIABLE (alphabetically), and finally by NUMBER
    processed_df["TYPE"] = processed_df["TYPE"].map(type_order)
    processed_df = processed_df.sort_values(by=["TYPE", "VARIABLE", "NUMBER"])

    # Convert back the TYPE from numerical values to the original categorical order
    processed_df["TYPE"] = processed_df["TYPE"].map(
        {v: k for k, v in type_order.items()}
    )

    output_sheets[sheet_name] = processed_df

# Save the file
savepath = filedialog.asksaveasfilename(
    defaultextension=".xlsx",
    filetypes=[("Excel files", "*.xlsx")],
    initialfile="Events_Synthesized.xlsx",
)

if savepath:
    with pd.ExcelWriter(savepath) as writer:
        for sheet_name, sheet_df in output_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
    messagebox.showinfo(
        "Save Successful",
        f"Events successfully compiled!",
    )
else:
    messagebox.showinfo(
        "Save Error", "Save operation canceled by user. Returning to main window."
    )
    sys.exit()
